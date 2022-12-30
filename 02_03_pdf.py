import csv
import itertools
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit

dic_currency = {'AZN':{'name':'Манаты','cost':35.68},'BYR':{'name':'Белорусские рубли','cost': 23.91},
    'EUR':{'name':'Евро','cost': 59.90},'GEL':{'name':'Грузинский лари','cost': 21.74},
    'KGS':{'name':'Киргизский сом','cost': 0.76},'KZT':{'name':'Тенге','cost': 0.13},
    'RUR':{'name':'Рубли','cost':1.00},'UAH':{'name':'Гривны','cost':  1.64},
    'USD':{'name':'Доллары','cost': 60.66},'UZS':{'name':'Узбекский сум','cost':0.0055}}

class Salary:
    def __init__(self, dict_vac):
        self.salary_from = dict_vac['salary_from']
        self.salary_to = dict_vac['salary_to']
        self.salary_currency = dict_vac['salary_currency']
    
class Vacancy:
    def __init__(self, dict_vac):
        self.name = dict_vac['name']
        self.salary = Salary(dict_vac)
        self.area_name = dict_vac['area_name']
        self.published_at = dict_vac['published_at']

class DataSet:
    @staticmethod    
    def reader_csv(filename):
        res = []
        with open(filename, encoding='utf-8-sig', newline='') as f:
            reader = csv.reader(f)
            for line in reader:
                if ('' in line): continue
                res.append(line)
        return res

    @staticmethod
    def _csv_ﬁler(res_data):
        if (len(res_data) == 0): return [{}]
        res_head = res_data.pop(0)
        return [dict(zip(res_head, row)) for row in res_data]
    
    @staticmethod
    def _csv_parser(task):
        res_data = DataSet.reader_csv(task.task_params['filename']['val'])
        ld = DataSet._csv_filer(res_data)
        list_vacancies = []
        for dct in ld:
            list_vacancies.append(Vacancy(dct))
        return list_vacancies

    def __init__(self, task):
        self.file_name = task.task_params['filename']['val']
        self.vac_obj = DataSet._csv_parser(task)
        self.dynamics_objects = DynamicObjects(task, self.vac_obj)

class DynamicObjects:  
    def __init__(self, task, vac_obj):
        self.salByYear     = {'name': 'Динамика уровня зарплат по годам', 'val': {}}                
        self.vacByYear     = {'name': 'Динамика количества вакансий по годам', 'val': {}} 
        self.salByYearProf = {'name': 'Динамика уровня зарплат по годам для выбранной профессии', 'val': {}}
        self.vacByYearProf = {'name': 'Динамика количества вакансий по годам для выбранной профессии', 'val': {}}
        self.salByCity     = {'name': 'Уровень зарплат по городам (в порядке убывания)', 'val': {}} 
        self.vacByCity     = {'name': 'Доля вакансий по городам (в порядке убывания)', 'val': {}}
        for vac in vac_obj:
            year = int(vac.published_at[0:4])
            sal_m = ((float(vac.salary.salary_to)+float(vac.salary.salary_from)) *
                    dic_currency[vac.salary.salary_currency]['cost'])
            if (year in self.salByYear['val'].keys()):
                self.salByYear['val'][year] += sal_m
                self.vacByYear['val'][year] += 1
            else:
                self.salByYear['val'][year] = sal_m
                self.vacByYear['val'][year] = 1
            if (task.task_params['prof']['val'] in vac.name):
                if (year in self.salByYearProf['val'].keys()):
                    self.salByYearProf['val'][year] += sal_m
                    self.vacByYearProf['val'][year] += 1
                else:
                    self.salByYearProf['val'][year] = sal_m
                    self.vacByYearProf['val'][year] = 1
            city = vac.area_name
            if (city in self.salByCity['val'].keys()):
                self.salByCity['val'][city] += sal_m
                self.vacByCity['val'][city] += 1
            else:
                self.salByCity['val'][city] = sal_m
                self.vacByCity['val'][city] = 1
        for k in self.salByYear['val'].keys(): 
            self.salByYear['val'][k] = int(self.salByYear['val'][k] / (self.vacByYear['val'][k] * 2))
        self.salByYear['val'] = dict(sorted(self.salByYear['val'].items(), key = lambda x: x[0]))
        self.vacByYear['val'] = dict(sorted(self.vacByYear['val'].items(), key = lambda x: x[0]))
        if len(self.vacByYearProf['val']) == 0: 
            self.salByYearProf['val'] = {2022: 0}
            self.vacByYearProf['val'] = {2022: 0}
        else:
            for k in self.vacByYearProf['val'].keys():
                self.salByYearProf['val'][k] = int(self.salByYearProf['val'][k] / (self.vacByYearProf['val'][k] * 2))
        self.salByYearProf['val'] = dict(sorted(self.salByYearProf['val'].items(), key = lambda x: x[0]))
        self.vacByYearProf['val'] = dict(sorted(self.vacByYearProf['val'].items(), key = lambda x: x[0]))
        
        for c in self.salByCity['val'].keys():
            self.salByCity['val'][c] = int(self.salByCity['val'][c] / (self.vacByCity['val'][c] * 2))
        self.vacByCity['val'] = dict(filter(lambda x: x[1] >= len(vac_obj) / 100, self.vacByCity['val'].items()))
        self.salByCity['val'] = dict(filter(lambda x: self.vacByCity['val'].__contains__(x[0])  , self.salByCity['val'].items()))
        self.salByCity['val'] = dict(sorted(self.salByCity['val'].items(), key = lambda x: x[1], reverse=True))
        self.vacByCity['val'] = dict(sorted(self.vacByCity['val'].items(), key = lambda x: (-x[1])))
        for c in self.vacByCity['val']:
            self.vacByCity['val'][c] = round(self.vacByCity['val'][c] / len(vac_obj), 4)
        self.salByCity['val'] = dict(itertools.islice(self.salByCity['val'].items(), 10))
        self.vacByCity['val'] = dict(itertools.islice(self.vacByCity['val'].items(), 10))    
       
class InputConnect:
    task_params = {
    'filename': {'prompt': 'Введите название файла',     'val': ''},
    'prof': {'prompt': 'Введите название профессии', 'val': ''} 
    }

    @classmethod
    def get_task(self):
        for tp in self.task_params.keys():
            self.task_params[tp]['val'] = ' '.join(input(f'{self.task_params[tp]["prompt"]}: ').split())
        return True
    
class Report:
    def __init__(self, dataset):
        self.salaries_year_level = dataset.dynamics_objects.salByYear['val']
        self.vacancies_year_count = dataset.dynamics_objects.vacByYear['val']
        self.selected_salary_year_level = dataset.dynamics_objects.salByYearProf['val']
        self.selected_vacancy_year_count = dataset.dynamics_objects.vacByYearProf['val']
        self.salaries_city_level = dataset.dynamics_objects.salByCity['val']
        self.vacancies_city_count = dataset.dynamics_objects.vacByCity['val']

    def generate_excel(self, prof):
        workbook_ = Workbook()
        stat_by_year = workbook_.worksheets[0]
        stat_by_year.title = "Cтатистика по годам"
        stats_by_city = workbook_.create_sheet("Cтатистика по городам")
        stat_by_year.append(["Год", "Средняя зарплата", f"Средняя зарплата - {prof}",
                              "Количество вакансий", f"Количество вакансий - {prof}"])
        for i, year in enumerate(self.salaries_year_level.keys(), 2):
            stat_by_year.cell(row=i, column=1, value=year)
            for j, dictionary in enumerate((self.salaries_year_level, self.vacancies_year_count,
                        self.selected_salary_year_level, self.selected_vacancy_year_count), 2):
                stat_by_year.cell(row=i, column=j, value=dictionary[year])
        stats_by_city.append(["Город", "Уровень зарплат", "", "Город", "Доля вакансий"])
        for i, city in enumerate(self.salaries_city_level.keys(), 2):
            stats_by_city.cell(row=i, column=1, value=city)
            stats_by_city.cell(row=i, column=2, value=self.salaries_city_level[city])
        for i, city in enumerate(self.vacancies_city_count.keys(), 2):
            stats_by_city.cell(row=i, column=4, value=city)
            stats_by_city.cell(row=i, column=5, value=self.vacancies_city_count[city]).number_format = '0.00%'
        self.wb_style(workbook_)
        workbook_.save('report.xlsx')

    @staticmethod
    def wb_style(wb):
        bold_font = Font(bold=True)
        thin = Side(border_style="thin", color="000000")
        outline = Border(top=thin, left=thin, right=thin, bottom=thin)
        for worksheet in wb.worksheets:
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value) if cell.value is not None else "") for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 3
            for cell in worksheet[1]:
                cell.font = bold_font
            for column in tuple(worksheet.columns):
                if column[1].value is None:
                    continue
                for cell in column:
                    cell.border = outline

    def generate_image(self, prof):
        fig1, ((f11, f12), (f21, f22)) = plt.subplots(2, 2, figsize=(12, 7.5), layout='constrained')
        self.generate_salByYear_graph(f11, prof)
        self.generate_vacByYear_graph(f12, prof)
        self.generate_salByCity_graph(f21)
        self.generate_vacByCity_graph(f22)
        plt.savefig('graph.png')

    def generate_salByYear_graph(self, f, prof):
        f_labels = self.salaries_year_level.keys()
        x = np.arange(len(f_labels))
        width = 0.35
        f.bar(x - width / 2, self.salaries_year_level.values(), width, label='Средняя з/п')
        f.bar(x + width / 2, self.selected_salary_year_level.values(), width, label=f'З/п {prof}')
        f.set_xticks(x, f_labels, fontsize=8, rotation=90, ha='right')
        f.set_title("Уровень зарплат по годам")
        f.yaxis.grid(True)
        f.legend(fontsize=8, loc='upper left')

    def generate_vacByYear_graph(self, f, prof):
        f_labels = self.vacancies_year_count.keys()
        x = np.arange(len(f_labels))
        width = 0.35
        f.bar(x - width / 2, self.vacancies_year_count.values(), width, label='Количество вакансий')
        f.bar(x + width / 2, self.selected_vacancy_year_count.values(), label=f'Количество вакансий {prof}')
        f.set_xticks(x, f_labels, fontsize=8, rotation=90, ha='right')
        f.set_title("Количество вакансий по годам")
        f.yaxis.grid(True)
        f.legend(fontsize=8, loc='upper left')

    def generate_salByCity_graph(self, f):
        f_labels = self.salaries_city_level.keys()
        y_pos = np.arange(len(f_labels))
        f.barh(y_pos, self.salaries_city_level.values(), align='center')
        f.set_yticks(y_pos, fontsize=8, labels=f_labels)
        f.invert_yaxis()
        f.xaxis.grid(True)
        f.set_title("Уровень зарплат по городам")

    def generate_vacByCity_graph(self, f):
        f_labels = list(self.vacancies_city_count.keys())
        values = list(self.vacancies_city_count.values())
        f_labels.append('Другие')
        values.append(1 - sum(values))
        f.pie(values, labels=f_labels, textprops={'fontsize': 8}, startangle=0, labeldistance=1.1,
            colors=['tab:orange','tab:green','tab:red','tab:purple','tab:brown','tab:pink',
                    'tab:gray','tab:olive','tab:cyan','tab:blue','tab:blue'])
        f.set_title("Доля вакансий по городам")

    def generate_pdf(self, prof):
        h1, h2, h3 = (["Год", "Средняя зарплата", f"Средняя зарплата - {prof}", "Количество вакансий",
            f"Количество вакансий - {prof}"], ["Город", "Уровень зарплат"], ["Город", "Доля вакансий"])
        r1 = list(map(lambda year: [year] + [dict[year] for dict in (self.salaries_year_level, self.vacancies_year_count,
            self.selected_salary_year_level, self.selected_vacancy_year_count)], self.salaries_year_level.keys()))
        r2 = list(map(lambda city: [city, self.salaries_city_level[city]], self.salaries_city_level.keys()))
        r3 = list(map(lambda city: [city, f'{round(self.vacancies_city_count[city]*100,2)}%'], self.vacancies_city_count.keys()))
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render(graph_name='graph.png',prof=prof,h1=h1,h2=h2,h3=h3,r1=r1,r2=r2,r3=r3)
        config = pdfkit.configuration(wkhtmltopdf= r'C:/Python/02_03_pdf\wkhtmltopdf/wkhtmltopdf.exe')
        options = {'enable-local-file-access': None}
        pdfkit.from_string(pdf_template, 'report.pdf', options=options, configuration=config)

my_task = InputConnect()
my_task.get_task()
my_data = DataSet(my_task)
if (my_data.vac_obj == None): exit()

print(f'{my_data.dynamics_objects.salByYear["name"]}: {my_data.dynamics_objects.salByYear["val"]}')
print(f'{my_data.dynamics_objects.vacByYear["name"]}: {my_data.dynamics_objects.vacByYear["val"]}')
print(f'{my_data.dynamics_objects.salByYearProf["name"]}: {my_data.dynamics_objects.salByYearProf["val"]}')
print(f'{my_data.dynamics_objects.vacByYearProf["name"]}: {my_data.dynamics_objects.vacByYearProf["val"]}')
print(f'{my_data.dynamics_objects.salByCity["name"]}: {my_data.dynamics_objects.salByCity["val"]}')
print(f'{my_data.dynamics_objects.vacByCity["name"]}: {my_data.dynamics_objects.vacByCity["val"]}')

report = Report(my_data)
report.generate_excel(my_task.task_params['prof']['val'])
report.generate_image(my_task.task_params['prof']['val'])
report.generate_pdf(my_task.task_params['prof']['val'])